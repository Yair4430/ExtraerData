import pandas as pd
from pathlib import Path
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
from typing import List
from .data_models import DocumentoData
from .constants import logger

class ExcelExporter:
    """Clase para exportar datos a Excel con formato y validaciones"""
    
    def __init__(self):
        pass
    
    def export_to_excel(self, extracted_data: List[DocumentoData], ficha: str) -> str:
        """Exporta los datos a un archivo Excel con formato y validaciones"""
        if not extracted_data:
            raise ValueError("No hay datos para exportar")

        try:
            # Preparar datos para DataFrame
            data_for_df = []
            for data in extracted_data:
                data_dict = {
                    'TIPO DE DOCUMENTO': data.tipo_documento,
                    'NUMERO DE DOCUMENTO': data.numero_documento,
                    'NOMBRES Y APELLIDOS': data.nombres_apellidos,
                    'DIA': data.dia,
                    'MES': data.mes.upper() if data.mes else '',  # Convertir a mayúsculas
                    'AÑO': data.año
                }
                data_for_df.append(data_dict)

            # Crear DataFrame
            df = pd.DataFrame(data_for_df)

            # Guardar automáticamente en Descargas
            filename = f'plantilla_{ficha}.xlsx'
            downloads_path = str(Path.home() / "Downloads")
            file_path = os.path.join(downloads_path, filename)

            # Guardar Excel inicial
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Datos', index=False)

            # Aplicar formato y validaciones
            self.ajustar_formato_excel(file_path)
            self.agregar_validaciones_excel(file_path)

            logger.info(f"Datos exportados exitosamente a: {file_path}")
            return file_path

        except Exception as e:
            logger.error(f"Error exportando a Excel: {e}")
            raise

    def ajustar_formato_excel(self, file_path: str):
        """Ajusta el formato del archivo Excel"""
        wb = load_workbook(file_path)
        ws = wb.active

        # Ajustar ancho de columnas
        column_widths = {
            'A': 20,  # TIPO DE DOCUMENTO
            'B': 20,  # NUMERO DE DOCUMENTO
            'C': 40,  # NOMBRES Y APELLIDOS (más ancho para nombres completos)
            'D': 10,  # DIA
            'E': 15,  # MES
            'F': 10   # AÑO
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Formato de encabezados
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        wb.save(file_path)

    def agregar_validaciones_excel(self, file_path: str):
        """Agrega validaciones de datos al archivo Excel"""
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Obtener el año actual
        año_actual = datetime.now().year
        
        # 1. Validación para TIPO DE DOCUMENTO (Columna A)
        dv_tipo_doc = DataValidation(
            type="list",
            formula1='"CC,TI,CE,PPT"',
            allow_blank=True
        )
        dv_tipo_doc.error = 'Debe seleccionar uno de los valores válidos: CC, TI, CE, PPT'
        dv_tipo_doc.errorTitle = 'Entrada inválida'
        dv_tipo_doc.prompt = 'Seleccione un tipo de documento válido: CC, TI, CE, PPT'
        dv_tipo_doc.promptTitle = 'Tipo de Documento'
        dv_tipo_doc.showErrorMessage = True
        dv_tipo_doc.errorStyle = 'stop'
        dv_tipo_doc.add('A2:A1000')
        ws.add_data_validation(dv_tipo_doc)
        
        # 2. Validación para DIA (Columna D)
        dv_dia = DataValidation(
            type="whole",
            operator="between",
            formula1=1,
            formula2=31,
            allow_blank=True
        )
        dv_dia.error = 'El día debe ser un número entre 1 y 31'
        dv_dia.errorTitle = 'Día inválido'
        dv_dia.prompt = 'Ingrese un día válido (1-31)'
        dv_dia.promptTitle = 'Día'
        dv_dia.showErrorMessage = True
        dv_dia.errorStyle = 'stop'
        dv_dia.add('D2:D1000')
        ws.add_data_validation(dv_dia)
        
        # 3. Validación para MES (Columna E)
        dv_mes = DataValidation(
            type="list",
            formula1='"ENERO,FEBRERO,MARZO,ABRIL,MAYO,JUNIO,JULIO,AGOSTO,SEPTIEMBRE,OCTUBRE,NOVIEMBRE,DICIEMBRE"',
            allow_blank=True
        )
        dv_mes.error = 'Debe seleccionar un mes válido en mayúsculas'
        dv_mes.errorTitle = 'Mes inválido'
        dv_mes.prompt = 'Seleccione un mes válido'
        dv_mes.promptTitle = 'Mes'
        dv_mes.showErrorMessage = True
        dv_mes.errorStyle = 'stop'
        dv_mes.add('E2:E1000')
        ws.add_data_validation(dv_mes)
        
        # 4. Validación para AÑO (Columna F)
        dv_año = DataValidation(
            type="whole",
            operator="between",
            formula1=1900,
            formula2=año_actual,
            allow_blank=True
        )
        dv_año.error = f'El año debe estar entre 1900 y {año_actual}'
        dv_año.errorTitle = 'Año inválido'
        dv_año.prompt = f'Ingrese un año válido (1900-{año_actual})'
        dv_año.promptTitle = 'Año'
        dv_año.showErrorMessage = True
        dv_año.errorStyle = 'stop'
        dv_año.add('F2:F1000')
        ws.add_data_validation(dv_año)
        
        wb.save(file_path)